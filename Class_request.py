from PySide2 import QtCore, QtGui, QtWidgets
from Class_filtr import *
from Class_for_request_directory import *
import os
from openpyxl import Workbook
import sqlite3


class class_request(QtWidgets.QWidget,class_filtr, class_for_request_directory):
    def __init__(self):
        super().__init__()


        
        self.btn_update      = QtWidgets.QPushButton('Обновить', self)
        self.btn_update.clicked.connect(self.write)

        self.to_excel        = QtWidgets.QPushButton('В Excel', self)
        self.to_excel.clicked.connect(self.write_excel)



        self.string_row = QtWidgets.QLineEdit(self)
        self.string_row.setText('')


        self.dict_table = {'table_1':None,'table_2':None,'table_3':None,'table_4':None}
        self.tab = QtWidgets.QTabWidget()
        self.write_table()

  
        self.grid = QtWidgets.QGridLayout()
        

        self.grid.addWidget(self.btn_update, 1,1)
        self.grid.addWidget(self.to_excel, 1,2)
        self.grid.addWidget(self.string_row, 1,3)
        self.grid.addWidget(self.tab, 2,1,1,14)

        self.setLayout(self.grid)



    def from_sql(self, table_name):
        cursor = sqlite3.connect("source.db").cursor()
        cursor.execute('SELECT * FROM '+table_name)
        return [[i[0] for i in cursor.description]] + cursor.fetchall()
        
            

        

    def write(self):
        self.dict_list = {'table_1':self.from_sql('table_1'),
                          'table_2':self.from_sql('table_2'),
                          'table_3':self.from_sql('table_3'),
                          'table_4':self.from_sql('table_4')}

        for i in self.dict_list:
            self.write_list(i, self.dict_list[i])
            self.dict_table[i].cellPressed.connect(self.cell_calculating)







    def write_excel(self):
        os.chdir(str(QtWidgets.QFileDialog.getExistingDirectory(self, "Выберите папку для сохранения")))
        wb = Workbook()

        
        #--------------------------------------------------------------------------------------------------------------------------------
        for i in self.dict_list:
            wb.create_sheet(i)
            sheet = wb[i]
            for row in range(len(self.dict_list[i])):
                
                shiht_row = row+1

                for column in range(len(self.dict_list[i][row])):
                    sheet.cell(row = shiht_row, column = column+1).value = self.dict_list[i][row][column]
            

            sheet.auto_filter.ref = "A1:AK100000"
            sheet.freeze_panes = sheet['A2']
            sheet.sheet_view.zoomScale = 70
        wb.remove(wb['Sheet'])
        #--------------------------------------------------------------------------------------------------------------------------------
            

        wb.save('excel.xlsx')
        print('excel.xlsx - done...')
            
        
 
